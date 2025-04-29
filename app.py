
import streamlit as st
import pandas as pd
import zipfile
import tempfile
import os
import fitz  # PyMuPDF
import re
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Extrator de Contas de Energia", layout="centered")
st.title("⚡ Extrator de Contas de Energia Elétrica (PDF ➜ Excel)")

uploaded_files = st.file_uploader("Envie arquivos PDF ou um ZIP com vários PDFs:", type=["pdf", "zip"], accept_multiple_files=True)

def extrair_dados_pdf(nome_arquivo, conteudo):
    texto = ""
    with fitz.open(stream=conteudo, filetype="pdf") as doc:
        for page in doc:
            texto += page.get_text()

    uc = Path(nome_arquivo).stem.split("UC")[-1].strip().split(" ")[0]

    nota = re.search(r"NOTA FISCAL Nº (\d+)", texto)
    data_emissao = re.search(r"DATA DE EMISSÃO: (\d{2}/\d{2}/\d{4})", texto)
    cnpj = re.search(r"CNPJ/CPF:\s*([\d./-]+)", texto)
    nome_titular = re.search(r"(?:\n|^)([A-Z\s]+LTDA|Ltda|ME)\s*CNPJ/CPF", texto)
    valor_total = re.search(r"TOTAL\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)", texto)

    dados = {
        "Número da Nota Fiscal": nota.group(1) if nota else "N/D",
        "Data de Emissão": data_emissao.group(1) if data_emissao else "N/D",
        "CNPJ": cnpj.group(1) if cnpj else "N/D",
        "Nome do Titular": nome_titular.group(1).strip() if nome_titular else "N/D",
        "Valor Total NF": float(valor_total.group(1).replace(".", "").replace(",", ".")) if valor_total else 0.0,
        "Fornecimento": []
    }

    padrao_linha = re.compile(r"((?:CONSUMO|INJEÇÃO|CONSUMO NÃO COMPENSADO)[^\n]+)")
    fornecimentos = padrao_linha.findall(texto)

    for linha in fornecimentos:
        fornecimento_nome = re.search(r"^(.*?kWh)", linha)
        nome = fornecimento_nome.group(1).strip() if fornecimento_nome else linha.strip()

        valores = re.findall(r"[-]?\d{1,3}(?:\.\d{3})*,\d{2}", linha)
        valor_total = float(valores[-1].replace(".", "").replace(",", ".")) if valores else 0.0

        icms = 0.0
        pos_percentual = [m.end() for m in re.finditer(r"%", linha)]
        for pos in pos_percentual:
            match = re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha[pos:])
            if match:
                icms = float(match.group().replace(".", "").replace(",", "."))
                break

        dados["Fornecimento"].append({
            "Fornecimento": nome,
            "Valor (R$)": valor_total,
            "ICMS (R$)": icms
        })

    return uc, dados

def processar_arquivos(uploaded_files):
    arquivos_pdf = []

    for uploaded in uploaded_files:
        if uploaded.type == "application/zip":
            with tempfile.TemporaryDirectory() as tmpdir:
                zip_path = os.path.join(tmpdir, "temp.zip")
                with open(zip_path, "wb") as f:
                    f.write(uploaded.read())
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(tmpdir)
                    for file_name in zip_ref.namelist():
                        if file_name.lower().endswith(".pdf"):
                            with open(os.path.join(tmpdir, file_name), "rb") as f:
                                arquivos_pdf.append((file_name, f.read()))
        else:
            arquivos_pdf.append((uploaded.name, uploaded.read()))
    return arquivos_pdf

def criar_planilha(dados_pdf):
    wb = Workbook()
    for uc, info in dados_pdf.items():
        ws = wb.create_sheet(title=uc)

        ws.append(["Número da Nota Fiscal", info["Número da Nota Fiscal"]])
        ws.append(["Data de Emissão", info["Data de Emissão"]])
        ws.append(["CNPJ", info["CNPJ"]])
        ws.append(["Nome do Titular", info["Nome do Titular"]])
        ws.append(["Valor Total NF", info["Valor Total NF"]])
        ws.append([])

        df = pd.DataFrame(info["Fornecimento"])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for row in ws.iter_rows(min_row=6, max_row=6, max_col=3):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="404040")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=7, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        ws.sheet_view.showGridLines = False

    wb.remove(wb["Sheet"])
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

if uploaded_files:
    arquivos = processar_arquivos(uploaded_files)
    dados_extraidos = {}

    for nome_arquivo, conteudo in arquivos:
        uc, dados = extrair_dados_pdf(nome_arquivo, conteudo)
        dados_extraidos[uc] = dados

    excel_bytes = criar_planilha(dados_extraidos)
    st.success("✅ Planilha gerada com sucesso!")
    st.download_button("📥 Baixar Excel", data=excel_bytes, file_name="Extrato_Contas_Energia.xlsx")
else:
    st.info("Envie arquivos PDF ou um .zip para iniciar.")
